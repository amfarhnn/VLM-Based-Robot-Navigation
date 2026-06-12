from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "thesis_tables.xlsx"
SOURCE_FILES = [
    ROOT / "chapter_1_introduction.md",
    ROOT / "literature_review.md",
    ROOT / "chapter_3_methodology.md",
    ROOT / "chapter_4_results_and_discussion.md",
]

TABLE_CAPTION_RE = re.compile(r"^\*\*Table\s+(\d+\.\d+):\s*(.+?)\*\*$")

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
ALTERNATE_FILL = PatternFill("solid", fgColor="F5F9FC")
WHITE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="1F1F1F")
BODY_FONT = Font(name="Calibri", size=11, color="1F1F1F")
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
THIN_GRAY = Side(style="thin", color="B7C9D6")
CELL_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def split_row(line: str) -> list[str]:
    return [cell.strip().replace("`", "") for cell in line.strip().strip("|").split("|")]


def extract_tables(path: Path) -> list[dict[str, object]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    tables: list[dict[str, object]] = []
    index = 0
    while index < len(lines):
        match = TABLE_CAPTION_RE.match(lines[index])
        if not match:
            index += 1
            continue

        number, title = match.groups()
        cursor = index + 1
        while cursor < len(lines) and not lines[cursor].strip().startswith("|"):
            cursor += 1

        raw_rows: list[str] = []
        while cursor < len(lines) and lines[cursor].strip().startswith("|"):
            raw_rows.append(lines[cursor])
            cursor += 1

        if len(raw_rows) >= 2:
            tables.append(
                {
                    "number": number,
                    "title": title,
                    "source": path.name,
                    "header": split_row(raw_rows[0]),
                    "rows": [split_row(row) for row in raw_rows[2:]],
                }
            )
        index = cursor
    return tables


def set_column_widths(worksheet, header: list[str], rows: list[list[str]]) -> None:
    for index, heading in enumerate(header, start=1):
        values = [heading] + [row[index - 1] if index - 1 < len(row) else "" for row in rows]
        longest = max(len(str(value)) for value in values)
        width = min(55, max(14, longest * 0.85 + 3))
        worksheet.column_dimensions[get_column_letter(index)].width = width


def add_table_sheet(workbook: Workbook, item: dict[str, object]) -> None:
    number = str(item["number"])
    title = str(item["title"])
    source = str(item["source"])
    header = list(item["header"])
    rows = [list(row) for row in item["rows"]]
    worksheet = workbook.create_sheet(f"Table {number}")
    final_column = get_column_letter(len(header))

    worksheet.merge_cells(f"A1:{final_column}1")
    title_cell = worksheet["A1"]
    title_cell.value = f"Table {number}: {title}"
    title_cell.fill = TITLE_FILL
    title_cell.font = WHITE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells(f"A2:{final_column}2")
    source_cell = worksheet["A2"]
    source_cell.value = f"Source Markdown: {source}. Copy rows 4 onward into the DOCX after the matching caption."
    source_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    source_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = 4
    for column, value in enumerate(header, start=1):
        cell = worksheet.cell(header_row, column, value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[header_row].height = 34

    for row_index, row in enumerate(rows, start=header_row + 1):
        for column in range(1, len(header) + 1):
            value = row[column - 1] if column - 1 < len(row) else ""
            cell = worksheet.cell(row_index, column, value)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if (row_index - header_row) % 2 == 0:
                cell.fill = ALTERNATE_FILL
        approximate_lines = max(
            1,
            max(
                len(str(row[column - 1])) // max(12, int(worksheet.column_dimensions[get_column_letter(column)].width))
                if column - 1 < len(row)
                else 1
                for column in range(1, len(header) + 1)
            )
            + 1,
        )
        worksheet.row_dimensions[row_index].height = min(90, max(24, approximate_lines * 15))

    final_row = header_row + len(rows)
    table = Table(displayName=f"Table_{number.replace('.', '_')}", ref=f"A{header_row}:{final_column}{final_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    set_column_widths(worksheet, header, rows)
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = "4:4"
    worksheet.print_area = f"A1:{final_column}{final_row}"


def build_contents(workbook: Workbook, tables: list[dict[str, object]]) -> None:
    worksheet = workbook.active
    worksheet.title = "Contents"
    worksheet.append(["Worksheet", "Table Caption", "Source Markdown"])

    for cell in worksheet[1]:
        cell.fill = TITLE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER

    for item in tables:
        number = str(item["number"])
        title = str(item["title"])
        row = worksheet.max_row + 1
        worksheet.cell(row, 1, f"Table {number}")
        worksheet.cell(row, 1).hyperlink = f"#'Table {number}'!A1"
        worksheet.cell(row, 1).font = LINK_FONT
        worksheet.cell(row, 2, f"Table {number}: {title}")
        worksheet.cell(row, 3, str(item["source"]))
        for cell in worksheet[row]:
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row % 2 == 1:
            for cell in worksheet[row]:
                cell.fill = ALTERNATE_FILL

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 70
    worksheet.column_dimensions["C"].width = 34
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:C{worksheet.max_row}"
    worksheet.sheet_view.showGridLines = False


def main() -> None:
    tables: list[dict[str, object]] = []
    for source in SOURCE_FILES:
        tables.extend(extract_tables(source))

    workbook = Workbook()
    build_contents(workbook, tables)
    for item in tables:
        add_table_sheet(workbook, item)

    workbook.save(OUTPUT_PATH)
    print(f"Built {OUTPUT_PATH} with {len(tables)} table worksheets")


if __name__ == "__main__":
    main()
